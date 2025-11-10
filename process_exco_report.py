#!/usr/bin/env python3
"""
Script to create EXCOs workbook template from HTML files and fill with recommendation data.
"""

import html.parser
import re
from functools import lru_cache
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
import sys
import re

REPORT_TEMPLATE_FILE = '/Users/satyasagardandela/Downloads/monthly Reporting/report 2025 (003).htm'
SII_TEMPLATE_FILE = '/Users/satyasagardandela/Downloads/monthly Reporting/SII.htm'
ADDITIONAL_HTML_FILE = '/Users/satyasagardandela/Downloads/monthly Reporting/Copy of 20251105_ECAG 2 Reporting.htm'
FORMAT_TEMPLATE_FILE = '/Users/satyasagardandela/Downloads/monthly Reporting/1.htm'
FORMAT_TEMPLATE_WORKBOOK = '/Users/satyasagardandela/Downloads/monthly Reporting/EXCOs Reporting October 2025.xlsx'

MANUAL_SII_IDS = {
    '2023-002_F07-A01',
    '2025-033_F03-A01',
    '2025-060_F02-A01',
    '2025-221_F01-A01',
    '2025-221_F02-A01',
    '2025-223_F02-A01'
}

MANUAL_ADD_TO_ECAG_IDS = set()

MANUAL_EXTRA_SII_IDS = MANUAL_SII_IDS

MANUAL_REMOVE_IDS = {
    '2025-032_F01-A01',
    '2024-078_F02-A02',
    '2025-048_F01-A02'
}

def sanitize_excel_value(value, is_date=False):
    """Remove illegal characters from Excel cell values. Optionally convert to date."""
    if value is None or pd.isna(value):
        return ''
    
    # If this is a date column, try to convert to datetime
    if is_date:
        try:
            if isinstance(value, str) and value.strip():
                date_value = pd.to_datetime(value)
                return date_value.date()
            elif isinstance(value, pd.Timestamp):
                return value.date()
            elif isinstance(value, datetime):
                return value.date()
        except:
            # If date conversion fails, return as string
            pass
    
    value_str = str(value)
    # Remove control characters except tab, newline, and carriage return
    # Excel allows: \t, \n, \r
    value_str = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value_str)
    value_str = _fix_common_spellings(value_str)
    return value_str


def _fix_common_spellings(text):
    """Apply known spelling corrections and restore missing umlauts."""
    if not text:
        return text

    def _replace_borse(match):
        original = match.group(0)
        upper = original.isupper()
        capitalized = original[0].isupper()
        replacement = 'BÖRSE' if upper else 'Börse' if capitalized else 'börse'
        return replacement

    def _replace_bohm(match):
        original = match.group(0)
        upper = original.isupper()
        capitalized = original[0].isupper()
        replacement = 'BÖHM' if upper else 'Böhm' if capitalized else 'böhm'
        return replacement

    text = re.sub(r'(?i)br\s*se', _replace_borse, text)
    text = re.sub(r'(?i)bhm', _replace_bohm, text)
    return text

def beautify_text(text):
    if text is None:
        return ''
    s = str(text).replace('\r', ' ').replace('\n', ' ')
    s = re.sub(r'\s{2,}', ' ', s)
    s = re.sub(r'(?:(?<=^)|(?<=[.!?]\s))\s*-\s+', ' • ', s)
    return s.strip()

MULTILINE_COLUMNS = {
    'Finding Title',
    'Finding Description',
    'Recommendation Details',
    'Management Response',
    'Risk',
    'Root cause',
    'Relevant Legal Entities',
    'Recommendation Owner - Legal Entity',
    'Recommendation Owner - Business Unit',
    'Board member'
}

def parse_html_table(html_file):
    """Parse HTML table and return as DataFrame."""
    with open(html_file, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    
    soup = BeautifulSoup(content, 'html.parser')
    tables = soup.find_all('table')
    
    if not tables:
        return None
    
    # Find the main table
    # For Word documents, look for MsoNormalTable
    table = None
    for t in tables:
        if 'MsoNormalTable' in str(t.get('class', [])):
            table = t
            break
    
    if not table:
        table = tables[0]  # Fallback to first table
    
    # Extract headers - look for rows with class containing xl67 or xl69 (header style) or MsoNormalTable
    headers = []
    header_row = None
    
    # Check if this is a Word document table (MsoNormalTable)
    is_word_table = 'MsoNormalTable' in str(table.get('class', []))
    
    for row in table.find_all('tr'):
        cells = row.find_all(['td', 'th'])
        if cells and len(cells) >= 2:
            # For Word tables, look for bold text in first row
            if is_word_table:
                first_cell = cells[0]
                # Check if first cell has bold text
                bold_text = first_cell.find('b')
                if bold_text:
                    headers = []
                    for cell in cells:
                        cell_text = ' '.join(cell.stripped_strings)
                        headers.append(cell_text)
                    header_row = row
                    break
            else:
                # For Excel tables, check if cells have header styling (class xl67 or xl69 or xl70)
                if len(cells) > 5:
                    cell_classes = ' '.join(str(cell.get('class', [])) for cell in cells)
                    if 'xl67' in cell_classes or 'xl69' in cell_classes or 'xl70' in cell_classes:
                        # Get text from first cell to verify it's a header
                        first_cell_text = cells[0].get_text(strip=True)
                        if first_cell_text and ('Finding' in first_cell_text or 'Audit' in first_cell_text or 
                                               'Recommendation' in first_cell_text or first_cell_text == 'Finding ID'):
                            headers = []
                            for cell in cells:
                                cell_text = cell.get_text(strip=True)
                                # Handle cases where text is split across spans
                                if not cell_text:
                                    # Try to get text from all child elements
                                    cell_text = ' '.join(cell.stripped_strings)
                                headers.append(cell_text)
                            header_row = row
                            break
    
    if not headers:
        # Fallback: use the first non-empty row as headers
        for row in table.find_all('tr'):
            cells = row.find_all(['td', 'th'])
            if not cells:
                continue
            candidate_headers = []
            for cell in cells:
                cell_text = ' '.join(cell.stripped_strings)
                candidate_headers.append(cell_text)
            if any(candidate_headers):
                headers = candidate_headers
                header_row = row
                break
        if not headers:
            print("Warning: Could not find headers in table")
            return None
    
    # Extract data rows
    rows = []
    header_row_found = False
    
    for row in table.find_all('tr'):
        # Skip until we find the header row
        if not header_row_found:
            if row == header_row:
                header_row_found = True
            continue
        
        cells = row.find_all(['td', 'th'])
        if cells and len(cells) >= len(headers):
            row_data = []
            for cell in cells[:len(headers)]:
                # Get all text, handling nested spans
                cell_text = ' '.join(cell.stripped_strings)
                row_data.append(cell_text)
            
            if any(row_data):  # Skip completely empty rows
                rows.append(row_data)
    
    # Ensure all rows have same length as headers
    if headers:
        max_cols = len(headers)
        rows = [row[:max_cols] + [''] * (max_cols - len(row)) if len(row) < max_cols else row[:max_cols] 
                for row in rows]
    
    if headers and rows:
        df = pd.DataFrame(rows, columns=headers)
        # Clean up column names - remove extra whitespace
        df.columns = [col.strip() for col in df.columns]
        return df
    
    return None


def _normalize_header_key(header):
    if header is None:
        return ''
    return re.sub(r'\s+', ' ', str(header)).strip().lower()


@lru_cache(maxsize=None)
def get_xlsx_template_styles():
    """Extract formatting metadata from the Excel template workbook."""
    metadata = {
        'column_widths_by_header': {},
        'column_widths_by_letter': {},
        'header_height': None,
        'data_height': None,
        'header_font': None,
        'data_font': None,
        'header_alignment': None,
        'data_alignment': None,
        'header_fill': None,
    }

    if not FORMAT_TEMPLATE_WORKBOOK:
        return metadata

    try:
        wb = load_workbook(FORMAT_TEMPLATE_WORKBOOK, data_only=True)
        if 'ECAG open Recos' not in wb.sheetnames:
            return metadata

        ws = wb['ECAG open Recos']

        # Column widths by header/letter
        header_row = ws[1] if ws.max_row >= 1 else []
        for cell in header_row:
            header_value = cell.value.strip() if isinstance(cell.value, str) else cell.value
            if header_value:
                width = ws.column_dimensions[cell.column_letter].width
                if width:
                    metadata['column_widths_by_header'][_normalize_header_key(header_value)] = width

        for letter, dim in ws.column_dimensions.items():
            if dim.width:
                metadata['column_widths_by_letter'][letter] = dim.width

        # Row heights
        header_dim = ws.row_dimensions.get(1)
        if header_dim and header_dim.height:
            metadata['header_height'] = header_dim.height

        data_height = None
        for idx in sorted(ws.row_dimensions):
            if idx > 1:
                dim = ws.row_dimensions[idx]
                if dim.height:
                    data_height = dim.height
                    break
        if data_height:
            metadata['data_height'] = data_height

        # Style samples
        sample_header = ws['A1']
        metadata['header_font'] = sample_header.font
        metadata['header_alignment'] = sample_header.alignment
        metadata['header_fill'] = sample_header.fill

        if ws.max_row >= 2:
            sample_data = ws['A2']
            metadata['data_font'] = sample_data.font
            metadata['data_alignment'] = sample_data.alignment
        else:
            metadata['data_font'] = sample_header.font
            metadata['data_alignment'] = sample_header.alignment

    except Exception as exc:
        print(f"Warning: Unable to read formatting from {FORMAT_TEMPLATE_WORKBOOK}: {exc}")

    return metadata


@lru_cache(maxsize=None)
def get_template_dimensions(html_file):
    widths = {}
    header_height = 32.0
    data_height = 54.0
    try:
        with open(html_file, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        soup = BeautifulSoup(content, 'html.parser')
        table = soup.find('table')
        if not table:
            raise ValueError('No table found in template')

        header_row = None
        data_row_height_found = False
        for row in table.find_all('tr'):
            cells = row.find_all(['td', 'th'])
            if not cells:
                continue

            style = row.get('style', '')
            height_pt = None
            match = re.search(r'height:([0-9.]+)pt', style)
            if match:
                height_pt = float(match.group(1))
            elif row.has_attr('height'):
                try:
                    height_pt = float(row['height'])
                except Exception:
                    height_pt = None

            if header_row is None:
                header_row = row
                if height_pt:
                    header_height = height_pt
                continue

            if not data_row_height_found and height_pt:
                data_height = height_pt
                data_row_height_found = True
                break

        if header_row is None:
            raise ValueError('No header row detected in template')

        cells = header_row.find_all(['td', 'th'])
        for cell in cells:
            header_text = ' '.join(cell.stripped_strings)
            header_text = re.sub(r'(?<=[a-z])\s+(?=[a-z])', '', header_text)
            key = _normalize_header_key(header_text)
            if not key:
                continue
            style = cell.get('style', '')
            width_pt = None
            match = re.search(r'width:([0-9.]+)pt', style)
            if match:
                width_pt = float(match.group(1))
            elif cell.get('width'):
                try:
                    width_pt = float(cell.get('width'))
                except Exception:
                    width_pt = None
            if width_pt is None:
                continue
            widths[key] = round(width_pt / 7.5, 2)
    except Exception as exc:
        print(f"Warning: Could not extract template dimensions from {html_file}: {exc}")
    return widths, header_height, data_height


def get_column_width_for_header(header, fallback, template_file):
    widths, _, _ = get_template_dimensions(template_file)
    key = _normalize_header_key(header)
    return widths.get(key, fallback)


def get_row_heights(template_file):
    _, header_height, data_height = get_template_dimensions(template_file)
    return header_height, data_height

def clean_dataframe(df):
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    df.columns = [re.sub(r'\s+', ' ', col.replace('\n', ' ')).strip() for col in df.columns]
    df = df.loc[:, [col for col in df.columns if col]]
    df = df.applymap(lambda x: re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', str(x)) if isinstance(x, str) else x)
    return df

def populate_board_member_flags(df):
    if df is None or df.empty or 'Board member' not in df.columns:
        return df

    df = df.copy()

    person_mapping = {
        'M. Graulich': ['MATTHIAS GRAULICH', 'GRAULICH'],
        'D. Senko': ['DANIEL SENKO', 'SENKO'],
        'J. Janka': ['JENS JANKA', 'JANKA'],
        'M. Matusza': ['MANFRED MATUSZA', 'MATUSZA']
    }

    def extract_entries(board_members_str):
        if pd.isna(board_members_str):
            return []
        cleaned = str(board_members_str).replace('\n', ' ')
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        if not cleaned:
            return []
        entries = re.findall(r'[^,]+?\([^)]*\)', cleaned)
        if not entries:
            entries = [segment.strip() for segment in cleaned.split(',') if segment.strip()]
        return entries

    def has_person(entry_list, variants):
        for entry in entry_list:
            entry_upper = entry.upper()
            for variant in variants:
                name_upper = variant.upper()
                if name_upper in entry_upper:
                    name_pos = entry_upper.find(name_upper)
                    paren_start = entry_upper.find('(', name_pos)
                    if paren_start >= 0:
                        paren_end = entry_upper.find(')', paren_start)
                        if paren_end >= 0:
                            entity = entry_upper[paren_start + 1:paren_end]
                            if 'EUREX CLEARING' in entity and 'EUREX REPO' not in entity:
                                return True
        return False

    entries_series = df['Board member'].apply(extract_entries)

    for column, variants in person_mapping.items():
        df[column] = entries_series.apply(lambda entries: 'Yes' if has_person(entries, variants) else 'No')

    return df

@lru_cache(maxsize=1)
def load_additional_html_df():
    try:
        df = parse_html_table(ADDITIONAL_HTML_FILE)
        if df is not None and not df.empty:
            df = clean_dataframe(df)
        return df if df is not None else pd.DataFrame()
    except Exception as exc:
        print(f"Warning: Could not load additional HTML file: {exc}")
        return pd.DataFrame()

def get_exco_headers():
    """Get the column headers for EXCO template based on the HTML structure."""
    return [
        'Finding ID',
        'Audit Year',
        'Audit ID',
        'Audit Title',
        'Recommendation ID',
        'Recommendation Type',
        'Recommendation Status',
        'Severity',
        'Finding Title',
        'Finding Description',
        'Recommendation Details',
        'Issue Date',
        'Initial Deadline',
        'Current Deadline',
        'No. of Deadline Shifts',
        'Relevant Legal Entities',
        'Recommendation Owner - Legal Entity',
        'Responsible Auditor',
        'Audit Lead',
        'Finding Owner - Responsible Board Member',
        'Recom Owner - Responsible Board Member',
        'Recommendation Owner - Division',
        'Recommendation Owner - Area',
        'Recommendation Owner - Managing Director',
        'Recommendation Owner - Business Unit',
        'Management Response',
        'Risk',
        'Recommendation Owner - Head of BU',
        'Board member',
        'Root cause',
        'Category',
        'Control Environment',
        'Control Design Effectiveness',
        'Control Operating Effectiveness',
        'M. Graulich',
        'D. Senko',
        'J. Janka',
        'M. Matusza'
    ]

def map_recommendations_to_exco(recommendations_df, finding_titles_df=None):
    """Map recommendations data to EXCO template format."""
    if recommendations_df is None or recommendations_df.empty:
        return pd.DataFrame()
    
    # Normalize column names (remove newlines and extra spaces)
    # Replace multiple spaces with single space, remove newlines
    recommendations_df.columns = [re.sub(r'\s+', ' ', col.replace('\n', ' ')).strip() for col in recommendations_df.columns]
    
    # Ensure we only work with valid Recommendation IDs
    if 'Recommendation ID' not in recommendations_df.columns:
        print("Warning: Recommendation ID column not found in recommendations file")
        return pd.DataFrame()
    
    # Remove any rows with empty or invalid Recommendation IDs
    recommendations_df = recommendations_df[recommendations_df['Recommendation ID'].astype(str).str.strip() != ''].copy()
    recommendations_df = recommendations_df[recommendations_df['Recommendation ID'].notna()].copy()
    
    # Create mapping dictionary - try to match with normalized column names
    def find_column(df, target_name):
        """Find column in dataframe that matches target name (case-insensitive, ignoring newlines and extra spaces)."""
        # Normalize: remove newlines, collapse multiple spaces, strip, lowercase
        target_normalized = re.sub(r'\s+', ' ', target_name.replace('\n', ' ')).strip().lower()
        for col in df.columns:
            col_normalized = re.sub(r'\s+', ' ', col.replace('\n', ' ')).strip().lower()
            if col_normalized == target_normalized:
                return col
        return None
    
    mapping = {}
    for exco_col in ['Finding ID', 'Audit Year', 'Audit ID', 'Audit Title', 'Recommendation ID',
                     'Recommendation Type', 'Recommendation Status', 'Severity', 'Finding Title',
                     'Finding Description', 'Recommendation Details', 'Issue Date', 'Initial Deadline',
                     'Current Deadline', 'No. of Deadline Shifts', 'Relevant Legal Entities',
                     'Recommendation Owner - Legal Entity', 'Responsible Auditor', 'Audit Lead',
                     'Finding Owner - Responsible Board Member', 'Recom Owner - Responsible Board Member',
                     'Recommendation Owner - Division', 'Recommendation Owner - Area',
                     'Recommendation Owner - Managing Director', 'Recommendation Owner - Business Unit',
                     'Management Response', 'Risk', 'Recommendation Owner - Head of BU',
                     'Board member', 'Root cause', 'Category']:
        found_col = find_column(recommendations_df, exco_col)
        if found_col:
            mapping[exco_col] = found_col
    
    # Create new dataframe with EXCO columns - initialize with proper length
    # IMPORTANT: Only use rows from recommendations_df - one row per Recommendation ID
    exco_headers = get_exco_headers()
    
    # Get unique Recommendation IDs to ensure we only have one row per ID
    if 'Recommendation ID' in recommendations_df.columns:
        # Keep only the first occurrence of each Recommendation ID
        recommendations_df = recommendations_df.drop_duplicates(subset=['Recommendation ID'], keep='first').copy()
    
    num_rows = len(recommendations_df)
    result_df = pd.DataFrame(index=range(num_rows), columns=exco_headers)
    
    # Map existing columns
    for exco_col in exco_headers:
        if exco_col in mapping:
            source_col = mapping[exco_col]
            if source_col in recommendations_df.columns:
                result_df[exco_col] = recommendations_df[source_col].values
            else:
                result_df[exco_col] = ''
        else:
            # For columns that don't have direct mapping, try to find them
            found_col = find_column(recommendations_df, exco_col)
            if found_col and found_col in recommendations_df.columns:
                result_df[exco_col] = recommendations_df[found_col].values
            else:
                result_df[exco_col] = ''
    
    # Debug: Print which columns were found
    print(f"Mapped columns: {list(mapping.keys())}")
    print(f"Available columns in recommendations: {[col for col in recommendations_df.columns if 'Owner' in col or 'Board' in col]}")
    
    # Handle Category column - split into Control Environment, Control Design Effectiveness, Control Operating Effectiveness
    category_col = find_column(recommendations_df, 'Category')
    if category_col and category_col in recommendations_df.columns:
        for idx, category in enumerate(recommendations_df[category_col]):
            if pd.notna(category):
                category_str = str(category)
                result_df.at[idx, 'Control Environment'] = 'Yes' if 'Control Env' in category_str else 'No'
                result_df.at[idx, 'Control Design Effectiveness'] = 'Yes' if 'Control Des' in category_str else 'No'
                result_df.at[idx, 'Control Operating Effectiveness'] = 'Yes' if 'Control Op' in category_str else 'No'
    
    # Fill the four person columns based on Board member column
    # Check if names appear with "Eurex Clearing AG" (not Eurex Repo)
    board_member_col = find_column(recommendations_df, 'Board member')
    if board_member_col and board_member_col in recommendations_df.columns:
        # Map names to columns - use full names and last names
        person_mapping = {
            'M. Graulich': ['MATTHIAS GRAULICH', 'GRAULICH'],
            'D. Senko': ['DANIEL SENKO', 'SENKO'],
            'J. Janka': ['JENS JANKA', 'JANKA'],
            'M. Matusza': ['MANFRED MATUSZA', 'MATUSZA']
        }
        
        for idx, board_members in enumerate(recommendations_df[board_member_col]):
            if pd.notna(board_members) and str(board_members).strip():
                board_members_str = str(board_members)
                
                # Check each person
                for person_col, name_variants in person_mapping.items():
                    found = False
                    
                    # Split by comma to handle multiple board members
                    # Format is typically: NAME1(Entity1), NAME2(Entity2), etc.
                    entries = board_members_str.split(',')
                    
                    for entry in entries:
                        entry_upper = entry.upper().strip()
                        
                        # Check each name variant
                        for name_variant in name_variants:
                            name_upper = name_variant.upper()
                            
                            # Check if this name appears in this entry
                            if name_upper in entry_upper:
                                # Check the pattern: NAME(Entity)
                                # We need to find NAME followed by (Eurex Clearing...) not (Eurex Repo...)
                                
                                # Look for the name followed by opening parenthesis
                                name_pos = entry_upper.find(name_upper)
                                if name_pos >= 0:
                                    # Find the opening parenthesis after the name
                                    paren_start = entry_upper.find('(', name_pos)
                                    if paren_start >= 0:
                                        # Extract the entity name in parentheses
                                        paren_end = entry_upper.find(')', paren_start)
                                        if paren_end >= 0:
                                            entity = entry_upper[paren_start + 1:paren_end]
                                            
                                            # Check if entity contains "Eurex Clearing" but NOT "Eurex Repo"
                                            if 'EUREX CLEARING' in entity and 'EUREX REPO' not in entity:
                                                found = True
                                                break
                        
                        if found:
                            break
                    
                    result_df.at[idx, person_col] = 'Yes' if found else 'No'
    
    # Complete Risk Description from Finding Title file if provided
    if finding_titles_df is not None and not finding_titles_df.empty:
        # Normalize Finding Title columns
        finding_title_col = None
        finding_desc_col = None
        
        for col in finding_titles_df.columns:
            col_normalized = col.replace('\n', ' ').strip().lower()
            if 'finding title' in col_normalized:
                finding_title_col = col
            elif 'finding description' in col_normalized:
                finding_desc_col = col
        
        # Try to match by Finding Title
        if finding_title_col and 'Finding Title' in result_df.columns:
            for idx, finding_title in enumerate(result_df['Finding Title']):
                if pd.notna(finding_title) and str(finding_title).strip():
                    current_risk = result_df.at[idx, 'Risk']
                    current_risk_str = str(current_risk) if pd.notna(current_risk) else ''
                    
                    # Check if risk description is incomplete (less than 50 chars or empty)
                    if len(current_risk_str.strip()) < 50:
                        # Try exact match first
                        matching_rows = finding_titles_df[
                            finding_titles_df[finding_title_col].str.strip().str.lower() == str(finding_title).strip().lower()
                        ]
                        
                        # If no exact match, try partial match
                        if matching_rows.empty:
                            title_part = str(finding_title)[:50].strip()
                            matching_rows = finding_titles_df[
                                finding_titles_df[finding_title_col].str.contains(title_part, case=False, na=False, regex=False)
                            ]
                        
                        if not matching_rows.empty:
                            # Use Finding Description as Risk if available
                            if finding_desc_col:
                                new_risk = matching_rows.iloc[0][finding_desc_col]
                                if pd.notna(new_risk) and len(str(new_risk).strip()) > len(current_risk_str.strip()):
                                    result_df.at[idx, 'Risk'] = new_risk
                            # Or try to find Risk column
                            elif 'Risk' in finding_titles_df.columns:
                                new_risk = matching_rows.iloc[0]['Risk']
                                if pd.notna(new_risk) and len(str(new_risk).strip()) > len(current_risk_str.strip()):
                                    result_df.at[idx, 'Risk'] = new_risk
    
    return result_df

def create_board_member_sheet(wb, df, board_member_name, board_member_col):
    """Create a sheet for a specific board member with their recommendations."""
    ws = wb.create_sheet(title=board_member_name)
    
    # Define styles
    header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # S3 - Yellow (matching report)
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # S4 - Blue
    pink_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Overdue - Pink
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # S1, S2 - White
    
    header_font = Font(bold=True, size=11, name="Calibri")
    normal_font = Font(size=11, name="Calibri")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Filter data for this board member
    board_member_df = df.copy()
    
    # Map board member names to search patterns
    name_mapping = {
        'M. Graulich': ['GRAULICH', 'MATTHIAS GRAULICH'],
        'D. Senko': ['SENKO', 'DANIEL SENKO'],
        'J. Janka': ['JANKA', 'JENS JANKA'],
        'M. Matusza': ['MATUSZA', 'MANFRED MATUSZA']
    }
    
    name_variants = name_mapping.get(board_member_name, [])
    
    # Filter rows where this board member appears with Eurex Clearing AG (not Eurex Repo)
    if board_member_col in board_member_df.columns:
        def matches_board_member(board_members_str):
            if pd.isna(board_members_str) or str(board_members_str).strip() == '':
                return False
            
            board_members_str = str(board_members_str)
            entries = board_members_str.split(',')
            
            for entry in entries:
                entry_upper = entry.upper().strip()
                
                for name_variant in name_variants:
                    name_upper = name_variant.upper()
                    
                    if name_upper in entry_upper:
                        # Check the pattern: NAME(Entity)
                        name_pos = entry_upper.find(name_upper)
                        if name_pos >= 0:
                            paren_start = entry_upper.find('(', name_pos)
                            if paren_start >= 0:
                                paren_end = entry_upper.find(')', paren_start)
                                if paren_end >= 0:
                                    entity = entry_upper[paren_start + 1:paren_end]
                                    
                                    # Check if entity contains "Eurex Clearing" but NOT "Eurex Repo"
                                    if 'EUREX CLEARING' in entity and 'EUREX REPO' not in entity:
                                        return True
            return False
        
        mask = board_member_df[board_member_col].apply(matches_board_member)
        board_member_df = board_member_df[mask].copy()
    
    if board_member_df.empty:
        ws.cell(row=1, column=1).value = f"No recommendations found for {board_member_name}"
        return
    
    # Sort by severity (same as main sheet)
    def get_sort_priority(row):
        severity = str(row.get('Severity', '')).strip().upper()
        current_deadline = row.get('Current Deadline', '')
        is_overdue = False
        today = pd.Timestamp.now().normalize()
        
        if pd.notna(current_deadline):
            try:
                if isinstance(current_deadline, str):
                    deadline_date = pd.to_datetime(current_deadline).normalize()
                else:
                    deadline_date = pd.to_datetime(current_deadline).normalize()
                if deadline_date <= today:
                    is_overdue = True
            except:
                pass
        
        recommendation_status = str(row.get('Recommendation Status', '')).strip()
        if is_overdue or (recommendation_status == 'Overdue'):
            return 0
        elif severity == 'S4':
            return 1
        elif severity == 'S3':
            return 2
        elif severity == 'S2':
            return 3
        elif severity == 'S1':
            return 4
        else:
            return 5
    
    board_member_df['_sort_priority'] = board_member_df.apply(get_sort_priority, axis=1)
    board_member_df = board_member_df.sort_values('_sort_priority', ascending=True).drop('_sort_priority', axis=1)
    board_member_df = board_member_df.reset_index(drop=True)
    
    # Write headers
    headers = get_exco_headers()
    date_columns = ['Issue Date', 'Initial Deadline', 'Current Deadline']
    number_columns = ['Audit Year', 'No. of Deadline Shifts']
    yes_no_columns = ['Control Environment', 'Control Design Effectiveness', 'Control Operating Effectiveness', 
                     'M. Graulich', 'D. Senko', 'J. Janka', 'M. Matusza']
    severity_column = 'Severity'
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
        # Match template alignment - headers left-aligned, top-vertical
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Write data
    for row_idx, (_, row) in enumerate(board_member_df.iterrows(), 2):
        severity = str(row.get('Severity', '')).strip().upper()
        recommendation_status = str(row.get('Recommendation Status', '')).strip()
        
        current_deadline = row.get('Current Deadline', '')
        is_overdue = False
        today = pd.Timestamp.now().normalize()
        if pd.notna(current_deadline):
            try:
                if isinstance(current_deadline, str):
                    deadline_date = pd.to_datetime(current_deadline).normalize()
                else:
                    deadline_date = pd.to_datetime(current_deadline).normalize()
                if deadline_date <= today:
                    is_overdue = True
            except:
                pass
        
        row_fill = None
        if is_overdue or (recommendation_status == 'Overdue'):
            row_fill = pink_fill
        elif severity == 'S4':
            row_fill = blue_fill
        elif severity == 'S3':
            row_fill = yellow_fill
        elif severity == 'S2' or severity == 'S1':
            row_fill = white_fill
        else:
            row_fill = white_fill
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row.get(header, '')
            # Check if this is a date column
            is_date_col = header in date_columns
            sanitized_value = sanitize_excel_value(value, is_date=is_date_col)
            if is_date_col:
                cell.value = sanitized_value
            else:
                if header in MULTILINE_COLUMNS and sanitized_value:
                    sanitized_value = beautify_text(sanitized_value)
                cell.value = sanitized_value
            
            # Apply date number format if it's a date column
            # Use "Short Date" format to match template and enable date filters
            if is_date_col and cell.value and cell.value != '':
                cell.number_format = 'dd.mm.yyyy'
                # Ensure date is stored as datetime object for proper filtering
                if not isinstance(cell.value, (datetime, pd.Timestamp)):
                    try:
                        cell.value = pd.to_datetime(cell.value)
                    except:
                        pass
            
            cell.font = normal_font
            cell.border = thin_border
            
            if header in date_columns or header in number_columns or header in yes_no_columns or header == severity_column:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_idx].height = data_row_height
    
    # Apply column widths based on template workbook where available
    column_widths_by_header = template_styles.get('column_widths_by_header', {})
    column_widths_by_letter = template_styles.get('column_widths_by_letter', {})
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        normalized = _normalize_header_key(header)
        width = column_widths_by_header.get(normalized)
        if width is None:
            width = column_widths_by_letter.get(col_letter)
        if width is not None:
            ws.column_dimensions[col_letter].width = width
        else:
            ws.column_dimensions[col_letter].width = 20
    
    ws.freeze_panes = 'A2'
    
    # Enable AutoFilter on all columns (for date filtering)
    headers = get_exco_headers()
    if len(headers) > 0:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col}1'
    
    return wb  # Return workbook for further sheet creation

def filter_board_member(df, board_member_name, board_member_col='Board member'):
    """Filter dataframe to include only recommendations for a specific board member (Eurex Clearing AG only)."""
    if df is None or df.empty or board_member_col not in df.columns:
        return pd.DataFrame()
    
    if board_member_name in df.columns:
        mask = df[board_member_name].astype(str).str.upper() == 'YES'
        return df[mask].copy()

    name_mapping = {
        'M. Graulich': ['GRAULICH', 'MATTHIAS GRAULICH'],
        'D. Senko': ['SENKO', 'DANIEL SENKO'],
        'J. Janka': ['JANKA', 'JENS JANKA'],
        'M. Matusza': ['MATUSZA', 'MANFRED MATUSZA']
    }

    name_variants = name_mapping.get(board_member_name, [])

    def matches_board_member(board_members_str):
        if pd.isna(board_members_str) or str(board_members_str).strip() == '':
            return False

        board_members_str_val = str(board_members_str)
        entries = board_members_str_val.split(',')

        for entry in entries:
            entry_upper = entry.upper().strip()
            for name_variant in name_variants:
                name_upper = name_variant.upper()
                if name_upper in entry_upper:
                    name_pos = entry_upper.find(name_upper)
                    if name_pos >= 0:
                        paren_start = entry_upper.find('(', name_pos)
                        if paren_start >= 0:
                            paren_end = entry_upper.find(')', paren_start)
                            if paren_end >= 0:
                                entity = entry_upper[paren_start + 1:paren_end]
                                if 'EUREX CLEARING' in entity and 'EUREX REPO' not in entity:
                                    return True
        return False

    mask = df[board_member_col].apply(matches_board_member)
    return df[mask].copy()

def write_filtered_data_to_sheet(wb, df, sheet_name, title=None, headers_override=None, template_file=FORMAT_TEMPLATE_FILE):
    """Write filtered dataframe to a new sheet with same formatting as main sheet."""
    if df is None or df.empty:
        return
    
    df = populate_board_member_flags(df)
    
    ws = wb.create_sheet(title=sheet_name)
    template_styles = get_xlsx_template_styles()
    
    # Define styles (same as main sheet)
    header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # S3 - Yellow (matching report)
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # S4 - Blue
    pink_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Overdue - Pink
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # S1, S2 - White
    
    header_font = Font(bold=True, size=11, name="Calibri")
    normal_font = Font(size=11, name="Calibri")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = headers_override if headers_override else get_exco_headers()
    date_columns = ['Issue Date', 'Initial Deadline', 'Current Deadline']
    number_columns = ['Audit Year', 'No. of Deadline Shifts']
    yes_no_columns = ['Control Environment', 'Control Design Effectiveness', 'Control Operating Effectiveness', 
                     'M. Graulich', 'D. Senko', 'J. Janka', 'M. Matusza']
    severity_column = 'Severity'
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
        # Match template alignment - headers left-aligned, top-vertical
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    header_row_height, template_data_row_height = get_row_heights(template_file)
    header_row_height = template_styles.get('header_height') or header_row_height or 42.0
    data_row_height = template_styles.get('data_height') or template_data_row_height or 54.0
    ws.sheet_format.defaultRowHeight = data_row_height
    ws.row_dimensions[1].height = header_row_height
    
    # Sort dataframe by severity (same logic as main sheet)
    def get_sort_priority(row):
        severity = str(row.get('Severity', '')).strip().upper()
        current_deadline = row.get('Current Deadline', '')
        is_overdue = False
        today = pd.Timestamp.now().normalize()
        
        if pd.notna(current_deadline):
            try:
                if isinstance(current_deadline, str):
                    deadline_date = pd.to_datetime(current_deadline).normalize()
                else:
                    deadline_date = pd.to_datetime(current_deadline).normalize()
                if deadline_date <= today:
                    is_overdue = True
            except:
                pass
        
        recommendation_status = str(row.get('Recommendation Status', '')).strip()
        if is_overdue or (recommendation_status == 'Overdue'):
            return 0
        elif severity == 'S4':
            return 1
        elif severity == 'S3':
            return 2
        elif severity == 'S2':
            return 3
        elif severity == 'S1':
            return 4
        else:
            return 5
    
    df_sorted = df.copy()
    df_sorted['_sort_priority'] = df_sorted.apply(get_sort_priority, axis=1)
    df_sorted = df_sorted.sort_values('_sort_priority', ascending=True).drop('_sort_priority', axis=1)
    df_sorted = df_sorted.reset_index(drop=True)

    # Write data
    for row_idx, (_, row) in enumerate(df_sorted.iterrows(), 2):
        severity = str(row.get('Severity', '')).strip().upper()
        recommendation_status = str(row.get('Recommendation Status', '')).strip()
        
        current_deadline = row.get('Current Deadline', '')
        is_overdue = False
        today = pd.Timestamp.now().normalize()
        if pd.notna(current_deadline):
            try:
                if isinstance(current_deadline, str):
                    deadline_date = pd.to_datetime(current_deadline).normalize()
                else:
                    deadline_date = pd.to_datetime(current_deadline).normalize()
                if deadline_date <= today:
                    is_overdue = True
            except:
                pass
        
        row_fill = None
        if is_overdue or (recommendation_status == 'Overdue'):
            row_fill = pink_fill
        elif severity == 'S4':
            row_fill = blue_fill
        elif severity == 'S3':
            row_fill = yellow_fill
        elif severity == 'S2' or severity == 'S1':
            row_fill = white_fill
        else:
            row_fill = white_fill
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row.get(header, '')
            # Check if this is a date column
            is_date_col = header in date_columns
            sanitized_value = sanitize_excel_value(value, is_date=is_date_col)
            if is_date_col:
                cell.value = sanitized_value
            else:
                if header in MULTILINE_COLUMNS and sanitized_value:
                    sanitized_value = beautify_text(sanitized_value)
                cell.value = sanitized_value
            
            # Apply date number format if it's a date column
            # Use "Short Date" format to match template and enable date filters
            if is_date_col and cell.value and cell.value != '':
                cell.number_format = 'dd.mm.yyyy'
                # Ensure date is stored as datetime object for proper filtering
                if not isinstance(cell.value, (datetime, pd.Timestamp)):
                    try:
                        cell.value = pd.to_datetime(cell.value)
                    except:
                        pass
            
            cell.font = normal_font
            cell.border = thin_border
            
            if header in date_columns or header in number_columns or header in yes_no_columns or header == severity_column:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_idx].height = data_row_height
    
    # Apply column widths aligned with template
    column_widths_by_header = template_styles.get('column_widths_by_header', {})
    column_widths_by_letter = template_styles.get('column_widths_by_letter', {})
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        normalized = _normalize_header_key(header)
        width = column_widths_by_header.get(normalized)
        if width is None:
            width = column_widths_by_letter.get(col_letter)
        if width is None:
            width = get_column_width_for_header(header, 20.0, template_file)
        ws.column_dimensions[col_letter].width = width
    
    ws.freeze_panes = 'A2'
    
    # Enable AutoFilter on all columns (for date filtering)
    if len(headers) > 0:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col}1'

def create_summary_sheet(wb, df):
    """Create a summary sheet with pivot tables for each board member."""
    ws = wb.create_sheet(title="Summary")
    
    header_font = Font(bold=True, size=12, name="Calibri")
    title_font = Font(bold=True, size=14, name="Calibri")
    normal_font = Font(size=11, name="Calibri")
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row = 1
    board_members = ['M. Graulich', 'D. Senko', 'J. Janka', 'M. Matusza']
    
    for board_member in board_members:
        # Title
        ws[f'A{row}'] = f"{board_member} - Summary"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        # Filter data for this board member
        board_member_df = filter_board_member(df, board_member)
        
        if not board_member_df.empty:
            # Create pivot table data
            total = len(board_member_df)
            
            # Severity breakdown
            severity_counts = board_member_df['Severity'].value_counts().to_dict()
            s4 = severity_counts.get('S4', 0)
            s3 = severity_counts.get('S3', 0)
            s2 = severity_counts.get('S2', 0)
            s1 = severity_counts.get('S1', 0)
            
            # Status breakdown
            status_counts = board_member_df['Recommendation Status'].value_counts().to_dict()
            
            # Overdue count
            overdue = 0
            if 'Current Deadline' in board_member_df.columns:
                for deadline in board_member_df['Current Deadline']:
                    if pd.notna(deadline):
                        try:
                            if isinstance(deadline, str):
                                deadline_date = pd.to_datetime(deadline).normalize()
                            else:
                                deadline_date = pd.to_datetime(deadline).normalize()
                            if deadline_date <= pd.Timestamp.now().normalize():
                                overdue += 1
                        except:
                            pass
            
            # Write headers
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            ws[f'C{row}'] = "Severity"
            ws[f'D{row}'] = "Count"
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Write total
            ws[f'A{row}'] = "Total Recommendations"
            ws[f'B{row}'] = total
            ws[f'B{row}'].font = Font(bold=True, size=11)
            ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Write severity breakdown
            ws[f'C{row}'] = "S4"
            ws[f'D{row}'] = s4
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            ws[f'C{row}'] = "S3"
            ws[f'D{row}'] = s3
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            ws[f'C{row}'] = "S2"
            ws[f'D{row}'] = s2
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            ws[f'C{row}'] = "S1"
            ws[f'D{row}'] = s1
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Write overdue
            ws[f'A{row}'] = "Overdue"
            ws[f'B{row}'] = overdue
            ws[f'B{row}'].font = Font(bold=True, size=11)
            ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 2
        else:
            ws[f'A{row}'] = "No recommendations found"
            row += 2
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def create_board_member_dashboard(wb, df, board_member_name, board_member_col):
    """Create a visual dashboard sheet for a specific board member."""
    ws = wb.create_sheet(title=f"{board_member_name} Dashboard")
    
    # Define styles
    title_font = Font(bold=True, size=18, name="Calibri")
    header_font = Font(bold=True, size=12, name="Calibri")
    normal_font = Font(size=11, name="Calibri")
    metric_font = Font(bold=True, size=14, name="Calibri")
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    metric_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Title
    ws['A1'] = f"{board_member_name} - Recommendations Dashboard"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    # Filter data for this board member
    board_member_df = df.copy()
    
    # Map board member names to search patterns
    name_mapping = {
        'M. Graulich': ['GRAULICH', 'MATTHIAS GRAULICH'],
        'D. Senko': ['SENKO', 'DANIEL SENKO'],
        'J. Janka': ['JANKA', 'JENS JANKA'],
        'M. Matusza': ['MATUSZA', 'MANFRED MATUSZA']
    }
    
    name_variants = name_mapping.get(board_member_name, [])
    
    # Filter rows where this board member appears with Eurex Clearing AG
    if board_member_col in board_member_df.columns:
        def matches_board_member(board_members_str):
            if pd.isna(board_members_str) or str(board_members_str).strip() == '':
                return False
            
            board_members_str_val = str(board_members_str)
            entries = board_members_str_val.split(',')
            
            for entry in entries:
                entry_upper = entry.upper().strip()
                for name_variant in name_variants:
                    name_upper = name_variant.upper()
                    if name_upper in entry_upper:
                        name_pos = entry_upper.find(name_upper)
                        if name_pos >= 0:
                            paren_start = entry_upper.find('(', name_pos)
                            if paren_start >= 0:
                                paren_end = entry_upper.find(')', paren_start)
                                if paren_end >= 0:
                                    entity = entry_upper[paren_start + 1:paren_end]
                                    if 'EUREX CLEARING' in entity and 'EUREX REPO' not in entity:
                                        return True
            return False
        
        mask = board_member_df[board_member_col].apply(matches_board_member)
        board_member_df = board_member_df[mask].copy()
    
    if board_member_df.empty:
        ws['A3'] = f"No recommendations found for {board_member_name}"
        return
    
    # Calculate key metrics for this board member
    total_recommendations = len(board_member_df)
    
    # Count by severity
    severity_counts = board_member_df['Severity'].value_counts().to_dict()
    s4_count = severity_counts.get('S4', 0)
    s3_count = severity_counts.get('S3', 0)
    s2_count = severity_counts.get('S2', 0)
    s1_count = severity_counts.get('S1', 0)
    
    # Count overdue
    overdue_count = 0
    if 'Current Deadline' in board_member_df.columns:
        for idx, deadline in enumerate(board_member_df['Current Deadline']):
            if pd.notna(deadline):
                try:
                    if isinstance(deadline, str):
                        deadline_date = pd.to_datetime(deadline).normalize()
                    else:
                        deadline_date = pd.to_datetime(deadline).normalize()
                    if deadline_date <= pd.Timestamp.now().normalize():
                        overdue_count += 1
                except:
                    pass
    
    # Write Key Metrics Section
    row = 3
    ws[f'A{row}'] = "Key Metrics"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    metrics = [
        ("Total Recommendations", total_recommendations),
        ("Overdue", overdue_count),
        ("S4 (Most Severe)", s4_count),
        ("S3", s3_count),
        ("S2", s2_count),
        ("S1 (Least Severe)", s1_count),
    ]
    
    for metric_name, metric_value in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = metric_value
        ws[f'B{row}'].font = metric_font
        ws[f'B{row}'].fill = metric_fill
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    # Create data for charts (starting from column D)
    chart_start_row = 3
    chart_data_row = chart_start_row + 1
    
    # Severity Distribution Chart Data
    ws['D3'] = "Severity Distribution"
    ws['D3'].font = header_font
    ws['D4'] = "Severity"
    ws['D4'].font = header_font
    ws['E4'] = "Count"
    ws['E4'].font = header_font
    
    severity_data = [
        ('S4', s4_count),
        ('S3', s3_count),
        ('S2', s2_count),
        ('S1', s1_count),
    ]
    
    for i, (sev, count) in enumerate(severity_data, start=5):
        ws[f'D{i}'] = sev
        ws[f'E{i}'] = count
    
    # Create Severity Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Recommendations by Severity"
    pie_chart.width = 10
    pie_chart.height = 7
    
    data = Reference(ws, min_col=5, min_row=4, max_row=8)
    cats = Reference(ws, min_col=4, min_row=5, max_row=8)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(cats)
    pie_chart.legend.position = 'r'
    ws.add_chart(pie_chart, 'D10')
    
    # Status Distribution Chart Data
    ws['H3'] = "Status Distribution"
    ws['H3'].font = header_font
    ws['H4'] = "Status"
    ws['H4'].font = header_font
    ws['I4'] = "Count"
    ws['I4'].font = header_font
    
    if 'Recommendation Status' in board_member_df.columns:
        status_counts = board_member_df['Recommendation Status'].value_counts().to_dict()
        status_data = list(status_counts.items())
        
        for i, (status, count) in enumerate(status_data, start=5):
            ws[f'H{i}'] = status
            ws[f'I{i}'] = count
        
        # Create Status Bar Chart
        if status_data:
            bar_chart = BarChart()
            bar_chart.title = "Recommendations by Status"
            bar_chart.type = "col"
            bar_chart.style = 10
            bar_chart.width = 10
            bar_chart.height = 7
            
            max_row = 4 + len(status_data)
            data = Reference(ws, min_col=9, min_row=4, max_row=max_row)
            cats = Reference(ws, min_col=8, min_row=5, max_row=max_row)
            bar_chart.add_data(data, titles_from_data=False)
            bar_chart.set_categories(cats)
            bar_chart.legend = None
            ws.add_chart(bar_chart, 'H10')
    
    # Category Distribution
    row += 2
    ws[f'A{row}'] = "Category Distribution"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    if 'Category' in board_member_df.columns:
        category_counts = board_member_df['Category'].value_counts().to_dict()
        for category, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
            ws[f'A{row}'] = str(category)[:40]  # Truncate long category names
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'] = count
            ws[f'B{row}'].font = metric_font
            ws[f'B{row}'].fill = metric_fill
            ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 10

def create_excel_workbook(df, output_file):
    """Create Excel workbook with formatting based on EXCO template."""
    additional_df = load_additional_html_df()

    if df is None or df.empty:
        df = pd.DataFrame(columns=get_exco_headers())

    # Manually add selected IDs from additional HTML source if missing
    if not additional_df.empty and MANUAL_ADD_TO_ECAG_IDS:
        addition_rows = additional_df[additional_df['Recommendation ID'].astype(str).str.strip().isin(MANUAL_ADD_TO_ECAG_IDS)].copy()
        if not addition_rows.empty:
            addition_rows = clean_dataframe(addition_rows)
            addition_rows = addition_rows.reindex(columns=get_exco_headers(), fill_value='')
            existing_ids = set(df['Recommendation ID'].astype(str).str.strip())
            addition_rows = addition_rows[~addition_rows['Recommendation ID'].astype(str).str.strip().isin(existing_ids)]
            if not addition_rows.empty:
                df = pd.concat([df, addition_rows], ignore_index=True)

    df = df.reindex(columns=get_exco_headers(), fill_value='')
    df['Recommendation ID'] = df['Recommendation ID'].astype(str).str.strip()
    df = df.drop_duplicates(subset=['Recommendation ID'], keep='first').reset_index(drop=True)

    if not df.empty and MANUAL_REMOVE_IDS:
        df = df[~df['Recommendation ID'].isin(MANUAL_REMOVE_IDS)].copy()

    df = populate_board_member_flags(df)

    # Keep original dataframe for SII sheets (before removing manual SII IDs)
    original_df = df.copy()

    if not df.empty and MANUAL_SII_IDS:
        df = df[~df['Recommendation ID'].isin(MANUAL_SII_IDS)].copy()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "ECAG open Recos"
    template_styles = get_xlsx_template_styles()
    
    # Define styles - match report color scheme
    header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # S3 - Yellow (matching report)
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # S4 - Blue
    pink_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Overdue - Pink
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # S1, S2 - White
    
    header_font = Font(bold=True, size=11, name="Calibri")
    normal_font = Font(size=11, name="Calibri")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = get_exco_headers()
    date_columns = ['Issue Date', 'Initial Deadline', 'Current Deadline']
    number_columns = ['Audit Year', 'No. of Deadline Shifts']
    yes_no_columns = ['Control Environment', 'Control Design Effectiveness', 'Control Operating Effectiveness', 
                     'M. Graulich', 'D. Senko', 'J. Janka', 'M. Matusza']
    severity_column = 'Severity'
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
        # Set header alignment - match template (left, top)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    header_row_height, template_data_row_height = get_row_heights(FORMAT_TEMPLATE_FILE)
    header_row_height = template_styles.get('header_height') or header_row_height or 42.0
    template_data_row_height = template_styles.get('data_height') or template_data_row_height or 90.75
    data_row_height = template_data_row_height
    ws.sheet_format.defaultRowHeight = data_row_height
    ws.row_dimensions[1].height = header_row_height
    
    # Filter out SII recommendations from main ECAG open Recos sheet
    # (SII recommendations are kept in separate sheets)
    if df is not None and not df.empty:
        if 'Recommendation Type' in df.columns:
            # Keep only non-SII recommendations for main sheet
            df = df[~df['Recommendation Type'].astype(str).str.contains('SII', case=False, na=False)].copy()
            print(f"Filtered out SII recommendations. Main sheet now has {len(df)} recommendations")
    
    # Sort dataframe by severity before writing
    # Order: S4 (most severe) -> S3 -> S2 -> S1 (least severe)
    # Overdue items should be first, then sorted by severity
    if df is not None and not df.empty:
        # Add a sort priority column
        def get_sort_priority(row):
            severity = str(row.get('Severity', '')).strip().upper()
            current_deadline = row.get('Current Deadline', '')
            is_overdue = False
            today = pd.Timestamp.now().normalize()
            
            # Check if overdue
            if pd.notna(current_deadline):
                try:
                    if isinstance(current_deadline, str):
                        deadline_date = pd.to_datetime(current_deadline).normalize()
                    else:
                        deadline_date = pd.to_datetime(current_deadline).normalize()
                    if deadline_date <= today:
                        is_overdue = True
                except:
                    pass
            
            recommendation_status = str(row.get('Recommendation Status', '')).strip()
            if is_overdue or (recommendation_status == 'Overdue'):
                # Overdue items get highest priority (0)
                return 0
            elif severity == 'S4':
                return 1
            elif severity == 'S3':
                return 2
            elif severity == 'S2':
                return 3
            elif severity == 'S1':
                return 4
            else:
                return 5  # Unknown severity goes last
        
        df['_sort_priority'] = df.apply(get_sort_priority, axis=1)
        df = df.sort_values('_sort_priority', ascending=True).drop('_sort_priority', axis=1)
        df = df.reset_index(drop=True)
    
    # Write data
    if df is not None and not df.empty:
        # Define which columns should be centered vs left-aligned
        date_columns = ['Issue Date', 'Initial Deadline', 'Current Deadline']
        number_columns = ['Audit Year', 'No. of Deadline Shifts']
        yes_no_columns = ['Control Environment', 'Control Design Effectiveness', 'Control Operating Effectiveness', 
                         'M. Graulich', 'D. Senko', 'J. Janka', 'M. Matusza']
        severity_column = 'Severity'
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            # Determine if this row should be colored (based on severity or overdue status)
            severity = str(row.get('Severity', '')).strip().upper()
            recommendation_status = str(row.get('Recommendation Status', '')).strip()
            
            # Check if overdue (pink)
            current_deadline = row.get('Current Deadline', '')
            is_overdue = False
            today = pd.Timestamp.now().normalize()
            if pd.notna(current_deadline):
                try:
                    if isinstance(current_deadline, str):
                        deadline_date = pd.to_datetime(current_deadline).normalize()
                    else:
                        deadline_date = pd.to_datetime(current_deadline).normalize()
                    if deadline_date <= today:
                        is_overdue = True
                except:
                    pass
            
            # Determine row color based on severity
            row_fill = None
            if is_overdue or (recommendation_status == 'Overdue'):
                # Overdue items are pink (highest priority)
                row_fill = pink_fill
            elif severity == 'S4':
                row_fill = blue_fill
            elif severity == 'S3':
                row_fill = yellow_fill
            elif severity == 'S2' or severity == 'S1':
                row_fill = white_fill
            else:
                # Unknown severity - white
                row_fill = white_fill
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row.get(header, '')
                # Check if this is a date column
                is_date_col = header in date_columns
                sanitized_value = sanitize_excel_value(value, is_date=is_date_col)
                if is_date_col:
                    cell.value = sanitized_value
                else:
                    if header in MULTILINE_COLUMNS and sanitized_value:
                        sanitized_value = beautify_text(sanitized_value)
                    cell.value = sanitized_value
                
                # Apply date number format if it's a date column
                # Use "Short Date" format to match template and enable date filters
                if is_date_col and cell.value and cell.value != '':
                    cell.number_format = 'dd.mm.yyyy'
                    # Ensure date is stored as datetime object for proper filtering
                    if not isinstance(cell.value, (datetime, pd.Timestamp)):
                        try:
                            cell.value = pd.to_datetime(cell.value)
                        except:
                            pass
                
                cell.font = normal_font
                cell.border = thin_border
                
                if header in date_columns or header in number_columns or header in yes_no_columns or header == severity_column:
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Apply color coding to each cell in the row
                if row_fill:
                    cell.fill = row_fill
            ws.row_dimensions[row_idx].height = data_row_height

    # Adjust column widths to mirror the template workbook
    column_widths_by_header = template_styles.get('column_widths_by_header', {})
    column_widths_by_letter = template_styles.get('column_widths_by_letter', {})
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        normalized = _normalize_header_key(header)
        width = column_widths_by_header.get(normalized)
        if width is None:
            width = column_widths_by_letter.get(col_letter)
        if width is not None:
            ws.column_dimensions[col_letter].width = width
        else:
            ws.column_dimensions[col_letter].width = get_column_width_for_header(header, 20.0, FORMAT_TEMPLATE_FILE)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Enable AutoFilter on all columns (for date filtering)
    if len(headers) > 0:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col}1'
    
    # Create individual recommendation sheets for each board member
    print("Creating board member recommendation sheets...")
    board_member_mapping = {
        'M. Graulich': 'MG Recommendations',
        'D. Senko': 'DS Recommendations',
        'J. Janka': 'JJ Recommendations',
        'M. Matusza': 'MM Recommendations'
    }
    
    for board_member, sheet_name in board_member_mapping.items():
        board_member_df = filter_board_member(df, board_member)
        if not board_member_df.empty:
            write_filtered_data_to_sheet(wb, board_member_df, sheet_name)
            print(f"  Created {sheet_name} ({len(board_member_df)} recommendations)")
    
    # Create SII recommendation sheets based on the uploaded SII template
    print("Creating SII recommendation sheets...")
    sii_df = parse_html_table(SII_TEMPLATE_FILE)
    if sii_df is not None and not sii_df.empty:
        # Normalize column names similar to other datasets
        sii_df.columns = [re.sub(r'\s+', ' ', col.replace('\n', ' ')).strip() for col in sii_df.columns]
        column_fixes = {
            'Finding type': 'Finding Type',
            'Recommendatio n Status': 'Recommendation Status',
            'Finding Owner - Responsible Boa rd Member': 'Finding Owner - Responsible Board Member',
            'Recom Owner - Responsible Boar d Member': 'Recom Owner - Responsible Board Member',
            'Recommendation Owner - Busin ess Unit': 'Recommendation Owner - Business Unit'
        }
        sii_df.rename(columns=lambda col: column_fixes.get(col, col), inplace=True)
        sii_df = populate_board_member_flags(sii_df)
        sii_headers = list(sii_df.columns)

        manual_sii_rows = original_df[original_df['Recommendation ID'].isin(MANUAL_SII_IDS)].copy()
        if not manual_sii_rows.empty:
            manual_sii_rows = manual_sii_rows.reindex(columns=sii_headers, fill_value='')
            manual_sii_rows = populate_board_member_flags(manual_sii_rows)

        remaining_manual_ids = MANUAL_SII_IDS - set(manual_sii_rows['Recommendation ID'].astype(str).str.strip())
        if remaining_manual_ids and not additional_df.empty:
            add_rows = additional_df[additional_df['Recommendation ID'].astype(str).str.strip().isin(remaining_manual_ids)].copy()
            if not add_rows.empty:
                add_rows = clean_dataframe(add_rows)
                add_rows = add_rows.reindex(columns=sii_headers, fill_value='')
                add_rows = populate_board_member_flags(add_rows)
                manual_sii_rows = pd.concat([manual_sii_rows, add_rows], ignore_index=True) if not manual_sii_rows.empty else add_rows

        if not manual_sii_rows.empty:
            sii_df = pd.concat([sii_df, manual_sii_rows], ignore_index=True)

        if 'Recommendation ID' in sii_df.columns:
            sii_df['Recommendation ID'] = sii_df['Recommendation ID'].astype(str).str.strip()
            sii_df = sii_df.drop_duplicates(subset=['Recommendation ID'], keep='first')
        write_filtered_data_to_sheet(wb, sii_df, 'SII Recommendations', headers_override=sii_headers, template_file=FORMAT_TEMPLATE_FILE)
        print(f"  Created SII Recommendations ({len(sii_df)} recommendations)")

        # SII ECAG (SII recommendations for ECAG/Eurex Clearing AG)
        sii_ecag_df = pd.DataFrame()
        if 'Relevant Legal Entities' in sii_df.columns:
            sii_ecag_df = sii_df[sii_df['Relevant Legal Entities'].astype(str).str.contains('ECAG|EUREX CLEARING', case=False, na=False, regex=True)].copy()
        if sii_ecag_df.empty and 'Recommendation Owner - Legal Entity' in sii_df.columns:
            sii_ecag_df = sii_df[sii_df['Recommendation Owner - Legal Entity'].astype(str).str.contains('ECAG|EUREX CLEARING', case=False, na=False, regex=True)].copy()
        if not sii_ecag_df.empty:
            write_filtered_data_to_sheet(wb, sii_ecag_df, 'SII ECAG', headers_override=sii_headers, template_file=FORMAT_TEMPLATE_FILE)
            print(f"  Created SII ECAG ({len(sii_ecag_df)} recommendations)")

        # SII recommendations for each board member
        for board_member, sheet_prefix in board_member_mapping.items():
            board_member_sii_df = filter_board_member(sii_df, board_member)
            if not board_member_sii_df.empty:
                sheet_name = f"{sheet_prefix.split()[0]} SII Recommendations"
                write_filtered_data_to_sheet(wb, board_member_sii_df, sheet_name, headers_override=sii_headers, template_file=FORMAT_TEMPLATE_FILE)
                print(f"  Created {sheet_name} ({len(board_member_sii_df)} recommendations)")
    
    # Create Notes sheet if there are recommendations with "Notes" in status
    print("Checking for Notes recommendations...")
    if 'Recommendation Status' in df.columns:
        notes_df = df[df['Recommendation Status'].astype(str).str.contains('Notes', case=False, na=False)].copy()
        if not notes_df.empty:
            write_filtered_data_to_sheet(wb, notes_df, 'Notes Recommendations')
            print(f"  Created Notes Recommendations ({len(notes_df)} recommendations)")
        else:
            print("  No Notes recommendations found")
    
    # Create Summary sheet with pivot tables (use original_df for complete summary)
    print("Creating Summary sheet...")
    create_summary_sheet(wb, original_df if original_df is not None and not original_df.empty else df)
    print("  Created Summary sheet")
    
    wb.save(output_file)
    print(f"Excel workbook created: {output_file}")

def main():
    # File paths
    exco_template_file = '/Users/satyasagardandela/Downloads/monthly Reporting/Copy of EXCOs Reporting October 2025 (003)1.htm'
    recommendations_file = '/Users/satyasagardandela/Downloads/monthly Reporting/Copy of Recommendations (10).htm'
    finding_title_file = '/Users/satyasagardandela/Downloads/monthly Reporting/Finding Title.htm'
    output_file = '/Users/satyasagardandela/Downloads/monthly Reporting/EXCO_Report_Output.xlsx'
    
    print("Parsing Recommendations file...")
    recommendations_df = parse_html_table(recommendations_file)
    
    if recommendations_df is None or recommendations_df.empty:
        print("Error: Could not parse recommendations file")
        return
    
    print(f"Found {len(recommendations_df)} recommendations")
    print(f"Columns: {list(recommendations_df.columns)}")
    
    # Try to parse Finding Title file
    finding_titles_df = None
    try:
        finding_titles_df = parse_html_table(finding_title_file)
        if finding_titles_df is not None:
            print(f"Found {len(finding_titles_df)} finding titles")
    except Exception as e:
        print(f"Note: Could not parse Finding Title file: {e}")
    
    print("Mapping data to EXCO template...")
    
    # CRITICAL: Ensure we only use Recommendation IDs from the recommendations file
    # Create a set of valid Recommendation IDs BEFORE mapping
    valid_recommendation_ids = set()
    if 'Recommendation ID' in recommendations_df.columns:
        valid_recommendation_ids = set(recommendations_df['Recommendation ID'].astype(str).str.strip())
        valid_recommendation_ids.discard('')  # Remove empty strings
        valid_recommendation_ids.update(MANUAL_ADD_TO_ECAG_IDS)
        print(f"Found {len(valid_recommendation_ids)} unique Recommendation IDs in recommendations file")
    
    # Map data
    exco_df = map_recommendations_to_exco(recommendations_df, finding_titles_df)
    
    # STRICT FILTER: Only keep rows where Recommendation ID exists in the recommendations file
    if 'Recommendation ID' in exco_df.columns and valid_recommendation_ids:
        exco_df['Recommendation ID'] = exco_df['Recommendation ID'].astype(str).str.strip()
        before_count = len(exco_df)
        
        # Get all IDs in exco_df before filtering
        exco_ids_before = set(exco_df['Recommendation ID'].unique())
        exco_ids_before.discard('')
        
        # Filter to ONLY valid IDs - use exact matching
        exco_df = exco_df[exco_df['Recommendation ID'].isin(valid_recommendation_ids)].copy()
        after_count = len(exco_df)
        
        # Get IDs after filtering
        exco_ids_after = set(exco_df['Recommendation ID'].unique())
        exco_ids_after.discard('')
        
        removed_count = before_count - after_count
        if removed_count > 0:
            print(f"WARNING: Removed {removed_count} rows with Recommendation IDs not in source file")
            removed_ids = exco_ids_before - valid_recommendation_ids
            print(f"Removed IDs: {sorted(removed_ids)}")
        else:
            print(f"All {after_count} recommendations have valid IDs from source file")
        
        # Final validation: Check for any mismatches
        invalid_ids = exco_ids_after - valid_recommendation_ids
        if invalid_ids:
            print(f"ERROR: Found {len(invalid_ids)} IDs in filtered data that are NOT in recommendations file:")
            for invalid_id in sorted(invalid_ids):
                print(f"  - {invalid_id}")
            # Remove any remaining invalid IDs
            exco_df = exco_df[exco_df['Recommendation ID'].isin(valid_recommendation_ids)].copy()
            print(f"Removed invalid IDs. Final count: {len(exco_df)}")
        else:
            print(f"✓ Validation passed: All {len(exco_df)} IDs match recommendations file")
    
    print("Creating Excel workbook...")
    create_excel_workbook(exco_df, output_file)
    
    # Final verification: Check the Excel file to ensure all IDs are valid
    print("\nPerforming final verification of Excel file...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb['ECAG open Recos']
        
        # Find Recommendation ID column
        rec_id_col = None
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(1, col).value or '').strip()
            if 'Recommendation ID' in header:
                rec_id_col = col
                break
        
        if rec_id_col:
            excel_ids = set()
            for row in range(2, ws.max_row + 1):
                rec_id = str(ws.cell(row, rec_id_col).value or '').strip()
                if rec_id:
                    excel_ids.add(rec_id)
            
            invalid_excel_ids = excel_ids - valid_recommendation_ids
            if invalid_excel_ids:
                print(f"ERROR: Found {len(invalid_excel_ids)} IDs in Excel file that are NOT in recommendations file:")
                for invalid_id in sorted(invalid_excel_ids):
                    print(f"  - {invalid_id}")
                print("This should not happen - please check the code.")
            else:
                print(f"✓ Final verification passed: All {len(excel_ids)} IDs in Excel file are from recommendations file")
        else:
            print("Could not find Recommendation ID column in Excel file for verification")
    except Exception as e:
        print(f"Note: Could not verify Excel file: {e}")
    
    print("Done!")

if __name__ == "__main__":
    main()

